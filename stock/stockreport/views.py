from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum
from datetime import datetime
from django.urls import reverse
from .models import master1, tran2, folio1, ClosingStock, ClosingBalance, APIConfig, UserAPIConfig, UserSession
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from django.utils.encoding import smart_str
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from io import BytesIO
from .services import LiveSQLService
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
import hashlib
import json


def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def generate_device_id(request):
    """Generate a unique device ID based on user agent and IP"""
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    ip = get_client_ip(request)
    device_string = f"{user_agent}:{ip}"
    return hashlib.md5(device_string.encode()).hexdigest()


def execute_query_user(query_type='tran2', user=None):
    """Execute SQL Server query using user-specific API configuration"""
    import pyodbc
    
    # Try to get user-specific config first
    if user:
        try:
            user_config = UserAPIConfig.objects.get(user=user, is_active=True)
            api_config = user_config
        except UserAPIConfig.DoesNotExist:
            # Fall back to global config
            api_config = APIConfig.get_active_config()
    else:
        # Fall back to global config
        api_config = APIConfig.get_active_config()
    
    if not api_config:
        raise ValueError("No active SQL Server configuration found")

    queries = {
        'tran2': "SELECT MasterCode1,VchType,RecType,VchNo,Date,Value1,Value3 FROM Tran2",
        'master1': "SELECT Code, MasterType, Name FROM Master1 WHERE MasterType IN (2,6,9)",
        'folio1': "SELECT MasterCode, MasterType, D1, D3 FROM folio1 WHERE MasterType in (2,6)"
    }
    query = queries[query_type]

    # Handle named instance (if port is blank) or host:port
    if api_config.port and api_config.port.strip():
        server = f"{api_config.url},{api_config.port}"
    else:
        server = api_config.url

    conn_str = (
        f"DRIVER={{ODBC Driver 11 for SQL Server}};"
        f"SERVER={server};"
        f"DATABASE={api_config.database};"
        f"UID={api_config.username};"
        f"PWD={api_config.password};"
        "TrustServerCertificate=yes;"
    )
    with pyodbc.connect(conn_str) as conn:
        cursor = conn.cursor()
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def user_signup(request):
    """User signup with API configuration"""
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            
            # Create user API configuration
            api_config = UserAPIConfig.objects.create(
                user=user,
                url=request.POST.get('api_url', ''),
                database=request.POST.get('api_database', 'your_db'),
                username=request.POST.get('api_username', ''),
                password=request.POST.get('api_password', ''),
                port=request.POST.get('api_port', ''),
                is_active=True
            )
            
            # Log the user in
            login(request, user)
            
            # Create session for single device login
            device_id = generate_device_id(request)
            UserSession.create_session(
                user=user,
                session_key=request.session.session_key,
                device_id=device_id,
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            messages.success(request, f'Account created successfully for {user.username}!')
            return redirect('home')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = UserCreationForm()
    
    return render(request, 'stockreport/signup.html', {'form': form})


def user_login(request):
    """User login with single device restriction"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Check if user has an active session on another device
            device_id = generate_device_id(request)
            active_sessions = UserSession.objects.filter(user=user, is_active=True)
            
            if active_sessions.exists():
                # Deactivate all existing sessions (single device login)
                active_sessions.update(is_active=False)
                messages.warning(request, 'You have been logged out from other devices.')
            
            # Log the user in
            login(request, user)
            
            # Create new session
            UserSession.create_session(
                user=user,
                session_key=request.session.session_key,
                device_id=device_id,
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            messages.success(request, f'Welcome back, {user.username}!')
            return redirect('home')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'stockreport/login.html')


def user_logout(request):
    """User logout"""
    if request.user.is_authenticated:
        # Deactivate user session
        UserSession.objects.filter(
            user=request.user,
            session_key=request.session.session_key
        ).update(is_active=False)
        
        logout(request)
        messages.success(request, 'You have been logged out successfully.')
    
    return redirect('login')


@login_required
def user_profile(request):
    """User profile and API configuration management"""
    try:
        api_config = request.user.api_config
    except UserAPIConfig.DoesNotExist:
        api_config = None
    
    if request.method == 'POST':
        # Update API configuration
        if api_config:
            api_config.url = request.POST.get('api_url', '')
            api_config.database = request.POST.get('api_database', 'your_db')
            api_config.username = request.POST.get('api_username', '')
            api_config.password = request.POST.get('api_password', '')
            api_config.port = request.POST.get('api_port', '')
            api_config.save()
        else:
            # Create new API configuration
            api_config = UserAPIConfig.objects.create(
                user=request.user,
                url=request.POST.get('api_url', ''),
                database=request.POST.get('api_database', 'your_db'),
                username=request.POST.get('api_username', ''),
                password=request.POST.get('api_password', ''),
                port=request.POST.get('api_port', ''),
                is_active=True
            )
        
        messages.success(request, 'API configuration updated successfully!')
        return redirect('user_profile')
    
    return render(request, 'stockreport/profile.html', {'api_config': api_config})


@login_required
def check_session_validity(request):
    """Check if current session is valid (for AJAX calls)"""
    if not request.user.is_authenticated:
        return HttpResponse(json.dumps({'valid': False, 'message': 'User not authenticated'}), 
                          content_type='application/json')
    
    device_id = generate_device_id(request)
    is_valid = UserSession.is_valid_session(
        user=request.user,
        session_key=request.session.session_key,
        device_id=device_id
    )
    
    if not is_valid:
        logout(request)
        return HttpResponse(json.dumps({'valid': False, 'message': 'Session expired'}), 
                          content_type='application/json')
    
    return HttpResponse(json.dumps({'valid': True}), content_type='application/json')


@login_required
def home(request):
    """Home page view with links to reports - requires login"""
    reports = [
        {
            'name': 'Closing Stock Report',
            'description': 'View closing stock for all inventory items with drill-down to detailed ledgers',
            'url': 'closing_stock_report'
        },
        {
            'name': 'Closing Balance Report',
            'description': 'View closing balance for all accounts',
            'url': 'closing_balance_report'
        }
    ]
    return render(request, 'stockreport/home.html', {'reports': reports})


@login_required
def closing_stock_report(request):
    """View for generating closing stock report with date filtering and data source selection
    
    Now uses Weighted Average Cost (WAC) for closing value calculation.
    Shows all items by default with option to filter zero quantity items.
    Supports both imported data and live SQL Server data.
    Uses user-specific API configurations.
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    data_source = request.GET.get('data_source', 'imported')  # 'imported' or 'live'
    hide_zero_balance = request.GET.get('hide_zero_balance', 'false').lower() == 'true'
    
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    results = []
    error_message = None
    
    try:
        if data_source == 'live':
            # Use live SQL Server data
            try:
                live_service = LiveSQLService(user=request.user)
                results = live_service.get_closing_stock_live(start_date_obj, end_date_obj, hide_zero_balance)
                # Ensure every item has an 'id' key for template compatibility
                for r in results:
                    if 'item' in r and 'code' in r['item']:
                        r['item']['id'] = r['item']['code']
            except ValueError as e:
                error_message = f"Live data error: {str(e)}"
                data_source = 'imported'  # Fallback to imported data
            except Exception as e:
                error_message = f"Error connecting to SQL Server: {str(e)}"
                data_source = 'imported'  # Fallback to imported data
        if data_source == 'imported' or not results:
            # Use imported data (existing logic)
            master_items = master1.objects.filter(mastertype=6).order_by('name')
            for item in master_items:
                # Get opening from folio1
                try:
                    folio_record = folio1.objects.get(MasterCode=item)
                    opening_qty = folio_record.D1 or 0
                    opening_val = folio_record.D3 or 0
                except folio1.DoesNotExist:
                    opening_qty = 0
                    opening_val = 0
                
                # Get all transactions for this item, sorted by date
                txns = tran2.objects.filter(mastercode1=item)
                if end_date_obj:
                    txns = txns.filter(date__lte=end_date_obj)
                txns = txns.order_by('date', 'id')
                
                # If start_date is set, calculate opening as of that date
                qty = opening_qty
                val = opening_val
                avg_rate = (val / qty) if qty else 0
                if start_date_obj:
                    for t in txns:
                        if t.date >= start_date_obj:
                            break
                        if t.quantity > 0:
                            # Purchase
                            total_cost = avg_rate * qty + t.amount
                            qty += t.quantity
                            val += t.amount
                            avg_rate = (val / qty) if qty else 0
                        elif t.quantity < 0:
                            # Sale
                            sale_qty = abs(t.quantity)
                            val -= sale_qty * avg_rate
                            qty -= sale_qty
                            avg_rate = (val / qty) if qty else 0
                    opening_qty = qty
                    opening_val = val
                
                # Now process transactions within the period (from start_date to end_date)
                period_qty = 0
                period_val = 0
                qty = opening_qty
                val = opening_val
                avg_rate = (val / qty) if qty else 0
                for t in txns:
                    if start_date_obj and t.date < start_date_obj:
                        continue
                    if t.quantity > 0:
                        # Purchase
                        qty += t.quantity
                        val += t.amount
                        avg_rate = (val / qty) if qty else 0
                        period_qty += t.quantity
                        period_val += t.amount
                    elif t.quantity < 0:
                        # Sale
                        sale_qty = abs(t.quantity)
                        val -= sale_qty * avg_rate
                        qty -= sale_qty
                        avg_rate = (val / qty) if qty else 0
                        period_qty += t.quantity
                        period_val -= sale_qty * avg_rate
                closing_qty = qty
                closing_val = val
                
                # Include all items, with option to filter zero quantity items
                if not hide_zero_balance or closing_qty != 0:
                    results.append({
                        'item': item,
                        'opening_quantity': round(opening_qty, 2),
                        'opening_value': round(opening_val, 2),
                        'transaction_quantity': round(period_qty, 2),
                        'transaction_value': round(period_val, 2),
                        'closing_quantity': round(closing_qty, 2),
                        'closing_value': round(closing_val, 2),
                    })
    except Exception as e:
        error_message = f"Error generating report: {str(e)}"
    
    # Calculate summary statistics
    total_items = len(results)
    zero_quantity_count = sum(1 for r in results if r['closing_quantity'] == 0)
    non_zero_quantity_count = total_items - zero_quantity_count
    
    # If no results and hide_zero_balance is True, show a message
    if total_items == 0 and hide_zero_balance:
        error_message = "No items found with non-zero closing quantity for the selected criteria."
    
    context = {
        'results': results,
        'start_date': start_date,
        'end_date': end_date,
        'data_source': data_source,
        'hide_zero_balance': hide_zero_balance,
        'error_message': error_message,
        'total_items': total_items,
        'zero_balance_count': zero_quantity_count,
        'non_zero_balance_count': non_zero_quantity_count,
    }
    return render(request, 'stockreport/closing_stock_report.html', context)


@login_required
def closing_balance_report(request):
    """View for generating closing balance report with date filtering and data source selection
    
    Filters out accounts with zero total value and sorts alphabetically.
    Supports both imported data and live SQL Server data.
    Uses user-specific API configurations.
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    data_source = request.GET.get('data_source', 'imported')  # 'imported' or 'live'
    
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    results = []
    error_message = None
    
    try:
        if data_source == 'live':
            # Use live SQL Server data
            try:
                live_service = LiveSQLService(user=request.user)
                results = live_service.get_closing_balance_live(start_date_obj, end_date_obj)
            except ValueError as e:
                error_message = f"Live data error: {str(e)}"
                data_source = 'imported'  # Fallback to imported data
            except Exception as e:
                error_message = f"Error connecting to SQL Server: {str(e)}"
                data_source = 'imported'  # Fallback to imported data
        else:
            # Use imported data (existing logic)
            master_items = master1.objects.filter(mastertype=2).order_by('name')
            for item in master_items:
                # Get base opening balance from folio1
                try:
                    folio_record = folio1.objects.get(MasterCode=item)
                    base_opening_balance = folio_record.D1 or 0
                except folio1.DoesNotExist:
                    base_opening_balance = 0
                
                # Calculate opening balance as of start_date (if provided)
                if start_date_obj:
                    # Get all transactions before start_date to calculate opening balance
                    opening_transactions = tran2.objects.filter(
                        mastercode1=item,
                        date__lt=start_date_obj
                    )
                    opening_amount = opening_transactions.aggregate(Sum('quantity'))['quantity__sum'] or 0
                    opening_balance = round(base_opening_balance + opening_amount, 2)
                else:
                    # Use folio1 opening balance if no start_date
                    opening_balance = base_opening_balance
                
                # Calculate closing balance up to end_date (if provided)
                if end_date_obj:
                    # Get all transactions up to end_date
                    closing_transactions = tran2.objects.filter(
                        mastercode1=item,
                        date__lte=end_date_obj
                    )
                    closing_amount = closing_transactions.aggregate(Sum('quantity'))['quantity__sum'] or 0
                    closing_balance = round(base_opening_balance + closing_amount, 2)
                else:
                    # Get all transactions if no end_date
                    closing_transactions = tran2.objects.filter(mastercode1=item)
                    closing_amount = closing_transactions.aggregate(Sum('quantity'))['quantity__sum'] or 0
                    closing_balance = round(base_opening_balance + closing_amount, 2)
                
                # Calculate transactions within the date range (for display purposes)
                period_transactions = tran2.objects.filter(mastercode1=item)
                if start_date_obj:
                    period_transactions = period_transactions.filter(date__gte=start_date_obj)
                if end_date_obj:
                    period_transactions = period_transactions.filter(date__lte=end_date_obj)
                
                period_amount = period_transactions.aggregate(Sum('quantity'))['quantity__sum'] or 0
                
                # Only include accounts with non-zero total value (opening + transactions + closing)
                total_value = abs(opening_balance) + abs(period_amount) + abs(closing_balance)
                if total_value > 0:
                    results.append({
                        'item': item,
                        'opening_balance': opening_balance,
                        'transaction_amount': period_amount,
                        'closing_balance': closing_balance
                    })
    
    except Exception as e:
        error_message = f"Error generating report: {str(e)}"
    
    context = {
        'results': results,
        'start_date': start_date,
        'end_date': end_date,
        'data_source': data_source,
        'error_message': error_message
    }
    
    return render(request, 'stockreport/closing_balance_report.html', context)


@login_required
def import_from_desktop(request):
    # Use user-specific import logic instead of admin redirect
    try:
        from django.contrib import messages
        from django.shortcuts import redirect
        from django.urls import reverse
        import time
        import datetime
        
        # Get user's API configuration
        try:
            user_config = UserAPIConfig.objects.get(user=request.user, is_active=True)
        except UserAPIConfig.DoesNotExist:
            messages.error(request, "No active SQL Server configuration found for your account. Please configure your API settings in your profile.")
            return redirect('home')
        
        # Step 1: Import Master1 data FIRST
        master_xml_response = execute_query_user(query_type='master1', user=request.user)
        if not master_xml_response:
            messages.error(request, "Failed to fetch Master1 data from external source.")
            return redirect('home')

        master_parsed_data = master_xml_response
        if not master_parsed_data:
            messages.error(request, "No Master1 data found or failed to parse XML response.")
            return redirect('home')

        old_master1_count = master1.objects.count()
        master1.objects.all().delete()
        messages.info(request, f"Cleared {old_master1_count} master1 records")

        master_imported_count = 0
        for entry in master_parsed_data:
            try:
                master1.objects.create(
                    code=int(entry['Code']),
                    name=entry['Name'],
                    mastertype=int(entry['MasterType'])
                )
                master_imported_count += 1
            except Exception as e:
                messages.warning(request, f"Error importing master1 record: {str(e)}")

        messages.success(request, f"Successfully imported {master_imported_count} master1 records")

        # Step 2: Import Tran2 data (AFTER master1)
        start_time = time.time()
        initial_tran2_count = tran2.objects.count()
        messages.info(request, f"Starting Tran2 import. Current database has {initial_tran2_count} records.")
        xml_response = execute_query_user(query_type='tran2', user=request.user)
        if not xml_response:
            messages.error(request, "Failed to fetch Tran2 data from external source.")
            return redirect('home')

        parsed_data = xml_response
        if not parsed_data:
            messages.error(request, "No Tran2 data found or failed to parse XML response.")
            return redirect('home')

        existing_records = {}
        for record in tran2.objects.select_related('mastercode1').all():
            key = (record.mastercode1.code, record.vch_no, record.date)
            existing_records[key] = record

        imported_count = 0
        updated_count = 0
        skipped_count = 0
        error_count = 0

        for entry in parsed_data:
            try:
                master_code = int(entry['MasterCode1'])
                if not master1.objects.filter(code=master_code).exists():
                    messages.warning(request, f"Skipped Tran2 record with non-existent master code {master_code} for vch_no {entry.get('VchNo', 'unknown')}")
                    error_count += 1
                    continue
                item = master1.objects.get(code=master_code)
                date_obj = entry['Date']
                if not date_obj:
                    messages.warning(request, f"Skipped Tran2 record with missing date for vch_no {entry.get('VchNo', 'unknown')}")
                    error_count += 1
                    continue
                if isinstance(date_obj, str):
                    try:
                        date_obj = datetime.datetime.strptime(date_obj, "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            date_obj = datetime.datetime.strptime(date_obj, "%Y-%m-%dT%H:%M:%S").date()
                        except ValueError:
                            messages.warning(request, f"Skipped Tran2 record with invalid date '{date_obj}' for vch_no {entry.get('VchNo', 'unknown')}")
                            error_count += 1
                            continue
                elif isinstance(date_obj, (datetime.datetime, datetime.date)):
                    date_obj = date_obj.date() if isinstance(date_obj, datetime.datetime) else date_obj
                else:
                    messages.warning(request, f"Skipped Tran2 record with unsupported date type for vch_no {entry.get('VchNo', 'unknown')}")
                    error_count += 1
                    continue
                new_rec_type = int(entry['RecType'])
                lookup_key = (item.code, entry['VchNo'].strip(), date_obj)
                existing_tran = existing_records.get(lookup_key)
                if existing_tran:
                    if new_rec_type > existing_tran.RecType:
                        existing_tran.Vch_Type = int(entry['VchType'])
                        existing_tran.RecType = new_rec_type
                        existing_tran.quantity = float(entry['Value1'])
                        existing_tran.amount = float(entry['Value3'])
                        existing_tran.save()
                        updated_count += 1
                    else:
                        skipped_count += 1
                else:
                    tran2.objects.create(
                        mastercode1=item,
                        Vch_Type=int(entry['VchType']),
                        RecType=new_rec_type,
                        vch_no=entry['VchNo'].strip(),
                        date=date_obj,
                        quantity=float(entry['Value1']),
                        amount=float(entry['Value3']),
                    )
                    imported_count += 1
            except Exception as e:
                error_count += 1
                messages.warning(request, f"Error importing Tran2 record: {str(e)}")

        elapsed_time = time.time() - start_time
        minutes = int(elapsed_time // 60)
        seconds = int(elapsed_time % 60)
        final_tran2_count = tran2.objects.count()
        total_change = final_tran2_count - initial_tran2_count
        messages.success(request, f"Tran2 import completed: {imported_count} created, {updated_count} updated, {skipped_count} skipped. Time taken: {minutes} min {seconds} sec. Database: {initial_tran2_count} → {final_tran2_count} records (+{total_change})")
        if error_count > 0:
            messages.warning(request, f"{error_count} Tran2 records had errors during import.")

        # Step 3: Import Folio1 data (AFTER master1)
        folio1_xml_response = execute_query_user(query_type='folio1', user=request.user)
        if not folio1_xml_response:
            messages.error(request, "Failed to fetch Folio1 data from external source.")
            return redirect('home')

        folio1_parsed_data = folio1_xml_response
        if not folio1_parsed_data:
            messages.error(request, "No Folio1 data found or failed to parse XML response.")
            return redirect('home')

        old_folio1_count = folio1.objects.count()
        folio1.objects.all().delete()
        messages.info(request, f"Cleared {old_folio1_count} folio1 records")

        folio1_imported_count = 0
        folio1_error_count = 0
        for entry in folio1_parsed_data:
            try:
                master_instance = master1.objects.get(code=int(entry['MasterCode']))
                folio1.objects.create(
                    MasterCode=master_instance,
                    D1=entry['D1'],
                    D3=entry['D3']
                )
                folio1_imported_count += 1
            except Exception as e:
                folio1_error_count += 1
                messages.warning(request, f"Error importing folio1 record: {str(e)}")

        messages.success(request, f"Successfully imported {folio1_imported_count} folio1 records")
        if folio1_error_count > 0:
            messages.warning(request, f"{folio1_error_count} folio1 records had errors during import.")

    except Exception as e:
        messages.error(request, f"Import failed: {str(e)}")
    
    return redirect('home')


@login_required
def export_closing_stock_excel(request):
    from datetime import datetime
    # Get date filters and data source
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    data_source = request.GET.get('data_source', 'imported')
    hide_zero_balance = request.GET.get('hide_zero_balance', 'false').lower() == 'true'
    
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Get data based on data source
    rows = []
    try:
        if data_source == 'live':
            # Use live SQL Server data
            live_service = LiveSQLService(user=request.user)
            results = live_service.get_closing_stock_live(start_date_obj, end_date_obj, hide_zero_balance)
            for result in results:
                rows.append([
                    result['item']['name'],
                    result['opening_quantity'],
                    result['opening_value'],
                    result['closing_quantity'],
                    result['closing_value'],
                ])
        else:
            # Use imported data (existing logic)
            master_items = master1.objects.filter(mastertype=6).order_by('name')
            for item in master_items:
                try:
                    folio_record = folio1.objects.get(MasterCode=item)
                    opening_qty = folio_record.D1 or 0
                    opening_val = folio_record.D3 or 0
                except folio1.DoesNotExist:
                    opening_qty = 0
                    opening_val = 0
                txns = tran2.objects.filter(mastercode1=item)
                if end_date_obj:
                    txns = txns.filter(date__lte=end_date_obj)
                txns = txns.order_by('date', 'id')
                qty = opening_qty
                val = opening_val
                avg_rate = (val / qty) if qty else 0
                if start_date_obj:
                    for t in txns:
                        if t.date >= start_date_obj:
                            break
                        if t.quantity > 0:
                            qty += t.quantity
                            val += t.amount
                            avg_rate = (val / qty) if qty else 0
                        elif t.quantity < 0:
                            sale_qty = abs(t.quantity)
                            val -= sale_qty * avg_rate
                            qty -= sale_qty
                            avg_rate = (val / qty) if qty else 0
                    opening_qty = qty
                    opening_val = val
                period_qty = 0
                period_val = 0
                qty = opening_qty
                val = opening_val
                avg_rate = (val / qty) if qty else 0
                for t in txns:
                    if start_date_obj and t.date < start_date_obj:
                        continue
                    if t.quantity > 0:
                        qty += t.quantity
                        val += t.amount
                        avg_rate = (val / qty) if qty else 0
                        period_qty += t.quantity
                        period_val += t.amount
                    elif t.quantity < 0:
                        sale_qty = abs(t.quantity)
                        val -= sale_qty * avg_rate
                        qty -= sale_qty
                        avg_rate = (val / qty) if qty else 0
                        period_qty += t.quantity
                        period_val -= sale_qty * avg_rate
                closing_qty = qty
                closing_val = val
                if not hide_zero_balance or closing_qty != 0:
                    rows.append([
                        item.name,
                        round(opening_qty, 2),
                        round(opening_val, 2),
                        round(closing_qty, 2),
                        round(closing_val, 2),
                    ])
    except Exception as e:
        # Return error response
        response = HttpResponse(f"Error generating Excel: {str(e)}", content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename=error.txt'
        return response
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Closing Stock Report'
    headers = ['Item Name', 'Opening Qty', 'Opening Value', 'Closing Qty', 'Closing Value']
    ws.append(headers)
    for row in rows:
        ws.append(row)
    # Set column widths
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=closing_stock_report.xlsx'
    wb.save(response)
    return response

@login_required
def export_closing_stock_pdf(request):
    """Export closing stock report to PDF"""
    from datetime import datetime
    # Get date filters and data source
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    data_source = request.GET.get('data_source', 'imported')
    hide_zero_balance = request.GET.get('hide_zero_balance', 'false').lower() == 'true'
    
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Get data based on data source
    rows = []
    try:
        if data_source == 'live':
            # Use live SQL Server data
            live_service = LiveSQLService(user=request.user)
            results = live_service.get_closing_stock_live(start_date_obj, end_date_obj, hide_zero_balance)
            for result in results:
                rows.append([
                    result['item']['name'],
                    result['opening_quantity'],
                    result['opening_value'],
                    result['closing_quantity'],
                    result['closing_value'],
                ])
        else:
            # Use imported data (existing logic)
            master_items = master1.objects.filter(mastertype=6).order_by('name')
            for item in master_items:
                try:
                    folio_record = folio1.objects.get(MasterCode=item)
                    opening_qty = folio_record.D1 or 0
                    opening_val = folio_record.D3 or 0
                except folio1.DoesNotExist:
                    opening_qty = 0
                    opening_val = 0
                txns = tran2.objects.filter(mastercode1=item)
                if end_date_obj:
                    txns = txns.filter(date__lte=end_date_obj)
                txns = txns.order_by('date', 'id')
                qty = opening_qty
                val = opening_val
                avg_rate = (val / qty) if qty else 0
                if start_date_obj:
                    for t in txns:
                        if t.date >= start_date_obj:
                            break
                        if t.quantity > 0:
                            qty += t.quantity
                            val += t.amount
                            avg_rate = (val / qty) if qty else 0
                        elif t.quantity < 0:
                            sale_qty = abs(t.quantity)
                            val -= sale_qty * avg_rate
                            qty -= sale_qty
                            avg_rate = (val / qty) if qty else 0
                    opening_qty = qty
                    opening_val = val
                period_qty = 0
                period_val = 0
                qty = opening_qty
                val = opening_val
                avg_rate = (val / qty) if qty else 0
                for t in txns:
                    if start_date_obj and t.date < start_date_obj:
                        continue
                    if t.quantity > 0:
                        qty += t.quantity
                        val += t.amount
                        avg_rate = (val / qty) if qty else 0
                        period_qty += t.quantity
                        period_val += t.amount
                    elif t.quantity < 0:
                        sale_qty = abs(t.quantity)
                        val -= sale_qty * avg_rate
                        qty -= sale_qty
                        avg_rate = (val / qty) if qty else 0
                        period_qty += t.quantity
                        period_val -= sale_qty * avg_rate
                closing_qty = qty
                closing_val = val
                if not hide_zero_balance or closing_qty != 0:
                    rows.append([
                        item.name,
                        round(opening_qty, 2),
                        round(opening_val, 2),
                        round(closing_qty, 2),
                        round(closing_val, 2),
                    ])
    except Exception as e:
        # Return error response
        response = HttpResponse(f"Error generating PDF: {str(e)}", content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename=error.txt'
        return response
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    title = Paragraph("Closing Stock Report", title_style)
    elements.append(title)
    
    # Add data source info
    data_source_text = f"Data Source: {'Live SQL Server' if data_source == 'live' else 'Imported Data'}"
    data_source_style = ParagraphStyle(
        'DataSource',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=10,
        alignment=1
    )
    data_source_para = Paragraph(data_source_text, data_source_style)
    elements.append(data_source_para)
    
    # Add date range if specified
    if start_date or end_date:
        date_text = "Date Range: "
        if start_date:
            date_text += f"From {start_date}"
        if end_date:
            date_text += f" To {end_date}"
        date_style = ParagraphStyle(
            'DateRange',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20,
            alignment=1
        )
        date_para = Paragraph(date_text, date_style)
        elements.append(date_para)
    
    # Add table
    if rows:
        headers = ['Item Name', 'Opening Qty', 'Opening Value', 'Closing Qty', 'Closing Value']
        table_data = [headers] + rows
        
        # Create table
        table = Table(table_data, colWidths=[2.5*inch, 1*inch, 1.2*inch, 1*inch, 1.2*inch])
        
        # Style the table
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),  # Right align numbers
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),    # Left align text
        ])
        table.setStyle(style)
        elements.append(table)
    else:
        no_data_style = ParagraphStyle(
            'NoData',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            alignment=1
        )
        no_data = Paragraph("No data available", no_data_style)
        elements.append(no_data)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=closing_stock_report.pdf'
    return response

@login_required
def export_closing_balance_excel(request):
    from datetime import datetime
    # Get date filters and data source
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    data_source = request.GET.get('data_source', 'imported')
    
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Get data based on data source
    rows = []
    try:
        if data_source == 'live':
            # Use live SQL Server data
            live_service = LiveSQLService(user=request.user)
            results = live_service.get_closing_balance_live(start_date_obj, end_date_obj)
            for result in results:
                rows.append([
                    result['item']['name'],
                    result['opening_balance'],
                    result['closing_balance'],
                ])
        else:
            # Use imported data (existing logic)
            master_items = master1.objects.filter(mastertype=2).order_by('name')
            for item in master_items:
                try:
                    folio_record = folio1.objects.get(MasterCode=item)
                    base_opening_balance = folio_record.D1 or 0
                except folio1.DoesNotExist:
                    base_opening_balance = 0
                if start_date_obj:
                    opening_transactions = tran2.objects.filter(
                        mastercode1=item,
                        date__lt=start_date_obj
                    )
                    opening_amount = opening_transactions.aggregate(Sum('quantity'))['quantity__sum'] or 0
                    opening_balance = round(base_opening_balance + opening_amount, 2)
                else:
                    opening_balance = base_opening_balance
                if end_date_obj:
                    closing_transactions = tran2.objects.filter(
                        mastercode1=item,
                        date__lte=end_date_obj
                    )
                    closing_amount = closing_transactions.aggregate(Sum('quantity'))['quantity__sum'] or 0
                    closing_balance = round(base_opening_balance + closing_amount, 2)
                else:
                    closing_transactions = tran2.objects.filter(mastercode1=item)
                    closing_amount = closing_transactions.aggregate(Sum('quantity'))['quantity__sum'] or 0
                    closing_balance = round(base_opening_balance + closing_amount, 2)
                period_transactions = tran2.objects.filter(mastercode1=item)
                if start_date_obj:
                    period_transactions = period_transactions.filter(date__gte=start_date_obj)
                if end_date_obj:
                    period_transactions = period_transactions.filter(date__lte=end_date_obj)
                period_amount = period_transactions.aggregate(Sum('quantity'))['quantity__sum'] or 0
                total_value = abs(opening_balance) + abs(period_amount) + abs(closing_balance)
                if total_value > 0:
                    rows.append([
                        item.name,
                        opening_balance,
                        closing_balance,
                    ])
    except Exception as e:
        # Return error response
        response = HttpResponse(f"Error generating Excel: {str(e)}", content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename=error.txt'
        return response
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Closing Balance Report'
    headers = ['Account Name', 'Opening Balance', 'Closing Balance']
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=closing_balance_report.xlsx'
    wb.save(response)
    return response

@login_required
def export_closing_balance_pdf(request):
    """Export closing balance report to PDF"""
    from datetime import datetime
    # Get date filters and data source
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    data_source = request.GET.get('data_source', 'imported')
    
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Get data based on data source
    rows = []
    try:
        if data_source == 'live':
            # Use live SQL Server data
            live_service = LiveSQLService(user=request.user)
            results = live_service.get_closing_balance_live(start_date_obj, end_date_obj)
            for result in results:
                rows.append([
                    result['item']['name'],
                    result['opening_balance'],
                    result['closing_balance'],
                ])
        else:
            # Use imported data (existing logic)
            master_items = master1.objects.filter(mastertype=2).order_by('name')
            for item in master_items:
                try:
                    folio_record = folio1.objects.get(MasterCode=item)
                    base_opening_balance = folio_record.D1 or 0
                except folio1.DoesNotExist:
                    base_opening_balance = 0
                if start_date_obj:
                    opening_transactions = tran2.objects.filter(
                        mastercode1=item,
                        date__lt=start_date_obj
                    )
                    opening_amount = opening_transactions.aggregate(Sum('quantity'))['quantity__sum'] or 0
                    opening_balance = round(base_opening_balance + opening_amount, 2)
                else:
                    opening_balance = base_opening_balance
                if end_date_obj:
                    closing_transactions = tran2.objects.filter(
                        mastercode1=item,
                        date__lte=end_date_obj
                    )
                    closing_amount = closing_transactions.aggregate(Sum('quantity'))['quantity__sum'] or 0
                    closing_balance = round(base_opening_balance + closing_amount, 2)
                else:
                    closing_transactions = tran2.objects.filter(mastercode1=item)
                    closing_amount = closing_transactions.aggregate(Sum('quantity'))['quantity__sum'] or 0
                    closing_balance = round(base_opening_balance + closing_amount, 2)
                period_transactions = tran2.objects.filter(mastercode1=item)
                if start_date_obj:
                    period_transactions = period_transactions.filter(date__gte=start_date_obj)
                if end_date_obj:
                    period_transactions = period_transactions.filter(date__lte=end_date_obj)
                period_amount = period_transactions.aggregate(Sum('quantity'))['quantity__sum'] or 0
                total_value = abs(opening_balance) + abs(period_amount) + abs(closing_balance)
                if total_value > 0:
                    rows.append([
                        item.name,
                        opening_balance,
                        closing_balance,
                    ])
    except Exception as e:
        # Return error response
        response = HttpResponse(f"Error generating PDF: {str(e)}", content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename=error.txt'
        return response
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    title = Paragraph("Closing Balance Report", title_style)
    elements.append(title)
    
    # Add data source info
    data_source_text = f"Data Source: {'Live SQL Server' if data_source == 'live' else 'Imported Data'}"
    data_source_style = ParagraphStyle(
        'DataSource',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=10,
        alignment=1
    )
    data_source_para = Paragraph(data_source_text, data_source_style)
    elements.append(data_source_para)
    
    # Add date range if specified
    if start_date or end_date:
        date_text = "Date Range: "
        if start_date:
            date_text += f"From {start_date}"
        if end_date:
            date_text += f" To {end_date}"
        date_style = ParagraphStyle(
            'DateRange',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20,
            alignment=1
        )
        date_para = Paragraph(date_text, date_style)
        elements.append(date_para)
    
    # Add table
    if rows:
        headers = ['Account Name', 'Opening Balance', 'Closing Balance']
        table_data = [headers] + rows
        
        # Create table
        table = Table(table_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        
        # Style the table
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),  # Right align numbers
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),    # Left align text
        ])
        table.setStyle(style)
        elements.append(table)
    else:
        no_data_style = ParagraphStyle(
            'NoData',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            alignment=1
        )
        no_data = Paragraph("No data available", no_data_style)
        elements.append(no_data)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=closing_balance_report.pdf'
    return response


@login_required
def stock_ledger_report(request, master_id):
    """
    Detailed stock ledger report for a specific item.
    Shows transaction history with running balances.
    Supports both imported data and live SQL Server data.
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    data_source = request.GET.get('data_source', 'imported')
    
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    ledger_data = []
    error_message = None
    master_item = None
    
    try:
        if data_source == 'live':
            live_service = LiveSQLService(user=request.user)
            # Fetch item details from SQL Server
            item_details = live_service.get_item_details(master_id)
            if not item_details:
                error_message = "Item not found in live SQL Server."
                master_item = {'name': f'Item {master_id}', 'code': master_id, 'mastertype': 6, 'id': master_id}
            else:
                master_item = item_details
            ledger_data = live_service.get_stock_ledger_live(master_id, start_date_obj, end_date_obj)
        else:
            master_item = get_object_or_404(master1, pk=master_id)
            # Use imported data
            try:
                folio_record = folio1.objects.get(MasterCode=master_item)
                opening_qty = folio_record.D1 or 0
                opening_val = folio_record.D3 or 0
            except folio1.DoesNotExist:
                opening_qty = 0
                opening_val = 0
            
            # Get all transactions for this item, sorted by date
            all_transactions = tran2.objects.filter(mastercode1=master_item).order_by('date', 'id')
            
            # Calculate opening balance as of start_date if provided
            if start_date_obj:
                qty = opening_qty
                val = opening_val
                avg_rate = (val / qty) if qty else 0
                
                for txn in all_transactions:
                    if txn.date >= start_date_obj:
                        break
                    if txn.quantity > 0:
                        # Purchase
                        qty += txn.quantity
                        val += txn.amount
                        avg_rate = (val / qty) if qty else 0
                    elif txn.quantity < 0:
                        # Sale
                        sale_qty = abs(txn.quantity)
                        val -= sale_qty * avg_rate
                        qty -= sale_qty
                        avg_rate = (val / qty) if qty else 0
                
                opening_qty = qty
                opening_val = val
            
            # Filter transactions for the period
            transactions = all_transactions
            if start_date_obj:
                transactions = transactions.filter(date__gte=start_date_obj)
            if end_date_obj:
                transactions = transactions.filter(date__lte=end_date_obj)
            
            # Calculate running balances
            running_qty = opening_qty
            running_val = opening_val
            avg_rate = (running_val / running_qty) if running_qty else 0
            
            # Add opening balance as first row if we have opening data
            if opening_qty > 0 or opening_val > 0:
                ledger_data.append({
                    'sno': 0,
                    'date': start_date_obj if start_date_obj else (transactions.first().date if transactions.exists() else datetime.now().date()),
                    'vchno': 'Opening Balance',
                    'opamount': 0,
                    'opqty': 0,
                    'qtyin': opening_qty,
                    'qtyout': 0,
                    'closingqty': opening_qty,
                    'closingamt': opening_val,
                    'description': f'Opening Balance as of {start_date_obj}' if start_date_obj else 'Opening Balance'
                })
            
            # Process each transaction
            for idx, txn in enumerate(transactions, 1):
                # Calculate opening values for this transaction
                op_qty = running_qty
                op_val = running_val
                
                # Determine qty in/out
                if txn.quantity > 0:
                    qty_in = txn.quantity
                    qty_out = 0
                    # Purchase - update running balances
                    running_qty += txn.quantity
                    running_val += txn.amount
                    avg_rate = (running_val / running_qty) if running_qty else 0
                else:
                    qty_in = 0
                    qty_out = abs(txn.quantity)
                    # Sale - update running balances
                    sale_qty = abs(txn.quantity)
                    running_val -= sale_qty * avg_rate
                    running_qty -= sale_qty
                
                ledger_data.append({
                    'sno': idx,
                    'date': txn.date,
                    'vchno': txn.vch_no,
                    'opamount': round(op_val, 2),
                    'opqty': round(op_qty, 2),
                    'qtyin': round(qty_in, 2),
                    'qtyout': round(qty_out, 2),
                    'closingqty': round(running_qty, 2),
                    'closingamt': round(running_val, 2),
                    'description': f"Voucher Type: {txn.Vch_Type}, RecType: {txn.RecType}"
                })
    
    except Exception as e:
        error_message = f"Error generating ledger: {str(e)}"
    
    # Calculate totals for template
    total_qtyin = sum(row['qtyin'] for row in ledger_data)
    total_qtyout = sum(row['qtyout'] for row in ledger_data)
    final_closingqty = ledger_data[-1]['closingqty'] if ledger_data else 0
    final_closingamt = ledger_data[-1]['closingamt'] if ledger_data else 0
    
    context = {
        'master_item': master_item,
        'ledger_data': ledger_data,
        'start_date': start_date,
        'end_date': end_date,
        'data_source': data_source,
        'error_message': error_message,
        'total_qtyin': total_qtyin,
        'total_qtyout': total_qtyout,
        'final_closingqty': final_closingqty,
        'final_closingamt': final_closingamt,
    }
    return render(request, 'stockreport/stock_ledger_report.html', context)


@login_required
def export_stock_ledger_excel(request, master_id):
    """Export stock ledger report to Excel"""
    master_item = get_object_or_404(master1, pk=master_id)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    data_source = request.GET.get('data_source', 'imported')
    
    # Get the same data as the view
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    ledger_data = []
    
    try:
        if data_source == 'live':
            try:
                live_service = LiveSQLService(user=request.user)
                ledger_data = live_service.get_stock_ledger_live(master_item.code, start_date_obj, end_date_obj)
            except:
                data_source = 'imported'
        
        if data_source == 'imported' or not ledger_data:
            # Use imported data (same logic as view)
            try:
                folio_record = folio1.objects.get(MasterCode=master_item)
                opening_qty = folio_record.D1 or 0
                opening_val = folio_record.D3 or 0
            except folio1.DoesNotExist:
                opening_qty = 0
                opening_val = 0
            
            transactions = tran2.objects.filter(mastercode1=master_item)
            if start_date_obj:
                transactions = transactions.filter(date__gte=start_date_obj)
            if end_date_obj:
                transactions = transactions.filter(date__lte=end_date_obj)
            transactions = transactions.order_by('date', 'id')
            
            running_qty = opening_qty
            running_val = opening_val
            avg_rate = (running_val / running_qty) if running_qty else 0
            
            if opening_qty > 0 or opening_val > 0:
                ledger_data.append({
                    'sno': 0,
                    'date': start_date_obj if start_date_obj else transactions.first().date if transactions.exists() else datetime.now().date(),
                    'vchno': 'Opening Balance',
                    'opamount': 0,
                    'opqty': 0,
                    'qtyin': opening_qty,
                    'qtyout': 0,
                    'closingqty': opening_qty,
                    'closingamt': opening_val,
                    'description': 'Opening Balance'
                })
            
            for idx, txn in enumerate(transactions, 1):
                op_qty = running_qty
                op_val = running_val
                
                if txn.quantity > 0:
                    qty_in = txn.quantity
                    qty_out = 0
                    running_qty += txn.quantity
                    running_val += txn.amount
                    avg_rate = (running_val / running_qty) if running_qty else 0
                else:
                    qty_in = 0
                    qty_out = abs(txn.quantity)
                    sale_qty = abs(txn.quantity)
                    running_val -= sale_qty * avg_rate
                    running_qty -= sale_qty
                
                ledger_data.append({
                    'sno': idx,
                    'date': txn.date,
                    'vchno': txn.vch_no,
                    'opamount': round(op_val, 2),
                    'opqty': round(op_qty, 2),
                    'qtyin': round(qty_in, 2),
                    'qtyout': round(qty_out, 2),
                    'closingqty': round(running_qty, 2),
                    'closingamt': round(running_val, 2),
                    'description': f"Voucher Type: {txn.Vch_Type}, RecType: {txn.RecType}"
                })
    except Exception as e:
        ledger_data = []
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Ledger"
    
    # Add title
    ws['A1'] = f"Stock Ledger Report - {master_item.name}"
    ws.merge_cells('A1:I1')
    ws['A1'].font = openpyxl.styles.Font(size=16, bold=True)
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # Add item details
    ws['A3'] = f"Item Name: {master_item.name}"
    ws['A4'] = f"Item Code: {master_item.code}"
    ws['A5'] = f"Master Type: {master_item.mastertype}"
    ws['A6'] = f"Data Source: {data_source.title()}"
    if start_date:
        ws['A7'] = f"Start Date: {start_date}"
    if end_date:
        ws['A8'] = f"End Date: {end_date}"
    
    # Add headers
    headers = ['S.No', 'Date', 'Voucher No', 'Opening Amount', 'Opening Qty', 'Qty In', 'Qty Out', 'Closing Qty', 'Closing Amount']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add data
    for row_idx, row_data in enumerate(ledger_data, 11):
        ws.cell(row=row_idx, column=1, value=row_data['sno'])
        ws.cell(row=row_idx, column=2, value=row_data['date'].strftime('%d/%m/%Y') if hasattr(row_data['date'], 'strftime') else str(row_data['date']))
        ws.cell(row=row_idx, column=3, value=row_data['vchno'])
        ws.cell(row=row_idx, column=4, value=row_data['opamount'])
        ws.cell(row=row_idx, column=5, value=row_data['opqty'])
        ws.cell(row=row_idx, column=6, value=row_data['qtyin'])
        ws.cell(row=row_idx, column=7, value=row_data['qtyout'])
        ws.cell(row=row_idx, column=8, value=row_data['closingqty'])
        ws.cell(row=row_idx, column=9, value=row_data['closingamt'])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=stock_ledger_{master_item.name}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response


@login_required
def export_stock_ledger_pdf(request, master_id):
    """Export stock ledger report to PDF"""
    master_item = get_object_or_404(master1, pk=master_id)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    data_source = request.GET.get('data_source', 'imported')
    
    # Get the same data as the view
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    ledger_data = []
    
    try:
        if data_source == 'live':
            try:
                live_service = LiveSQLService(user=request.user)
                ledger_data = live_service.get_stock_ledger_live(master_item.code, start_date_obj, end_date_obj)
            except:
                data_source = 'imported'
        
        if data_source == 'imported' or not ledger_data:
            # Use imported data (same logic as view)
            try:
                folio_record = folio1.objects.get(MasterCode=master_item)
                opening_qty = folio_record.D1 or 0
                opening_val = folio_record.D3 or 0
            except folio1.DoesNotExist:
                opening_qty = 0
                opening_val = 0
            
            transactions = tran2.objects.filter(mastercode1=master_item)
            if start_date_obj:
                transactions = transactions.filter(date__gte=start_date_obj)
            if end_date_obj:
                transactions = transactions.filter(date__lte=end_date_obj)
            transactions = transactions.order_by('date', 'id')
            
            running_qty = opening_qty
            running_val = opening_val
            avg_rate = (running_val / running_qty) if running_qty else 0
            
            if opening_qty > 0 or opening_val > 0:
                ledger_data.append({
                    'sno': 0,
                    'date': start_date_obj if start_date_obj else transactions.first().date if transactions.exists() else datetime.now().date(),
                    'vchno': 'Opening Balance',
                    'opamount': 0,
                    'opqty': 0,
                    'qtyin': opening_qty,
                    'qtyout': 0,
                    'closingqty': opening_qty,
                    'closingamt': opening_val,
                    'description': 'Opening Balance'
                })
            
            for idx, txn in enumerate(transactions, 1):
                op_qty = running_qty
                op_val = running_val
                
                if txn.quantity > 0:
                    qty_in = txn.quantity
                    qty_out = 0
                    running_qty += txn.quantity
                    running_val += txn.amount
                    avg_rate = (running_val / running_qty) if running_qty else 0
                else:
                    qty_in = 0
                    qty_out = abs(txn.quantity)
                    sale_qty = abs(txn.quantity)
                    running_val -= sale_qty * avg_rate
                    running_qty -= sale_qty
                
                ledger_data.append({
                    'sno': idx,
                    'date': txn.date,
                    'vchno': txn.vch_no,
                    'opamount': round(op_val, 2),
                    'opqty': round(op_qty, 2),
                    'qtyin': round(qty_in, 2),
                    'qtyout': round(qty_out, 2),
                    'closingqty': round(running_qty, 2),
                    'closingamt': round(running_val, 2),
                    'description': f"Voucher Type: {txn.Vch_Type}, RecType: {txn.RecType}"
                })
    except Exception as e:
        ledger_data = []
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    title = Paragraph(f"Stock Ledger Report - {master_item.name}", title_style)
    elements.append(title)
    
    # Item details
    details_data = [
        ['Item Name:', master_item.name],
        ['Item Code:', str(master_item.code)],
        ['Master Type:', str(master_item.mastertype)],
        ['Data Source:', data_source.title()],
    ]
    if start_date:
        details_data.append(['Start Date:', start_date])
    if end_date:
        details_data.append(['End Date:', end_date])
    
    details_table = Table(details_data, colWidths=[2*inch, 4*inch])
    details_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(details_table)
    elements.append(Spacer(1, 20))
    
    # Table headers
    headers = ['S.No', 'Date', 'Voucher No', 'Opening Amount', 'Opening Qty', 'Qty In', 'Qty Out', 'Closing Qty', 'Closing Amount']
    
    # Prepare table data
    table_data = [headers]
    for row_data in ledger_data:
        table_data.append([
            str(row_data['sno']),
            row_data['date'].strftime('%d/%m/%Y') if hasattr(row_data['date'], 'strftime') else str(row_data['date']),
            str(row_data['vchno']),
            str(row_data['opamount']),
            str(row_data['opqty']),
            str(row_data['qtyin']),
            str(row_data['qtyout']),
            str(row_data['closingqty']),
            str(row_data['closingamt'])
        ])
    
    if not ledger_data:
        table_data.append(['No transactions found for this item', '', '', '', '', '', '', '', ''])
    
    # Create table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=stock_ledger_{master_item.name}_{datetime.now().strftime("%Y%m%d")}.pdf'
    response.write(pdf)
    
    return response
